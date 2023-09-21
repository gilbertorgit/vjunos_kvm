"""
---------------------------------
 Author: gilbertorgit
 Date: 01/02/2023
---------------------------------
"""

import os
import openpyxl

"""
You can use this script to change the version easily
Specify the TAB -> tab_name
Specify the old value -> old_value
Specify the new value -> new_value

and run the script
"""

directory = './'

# The new value to replace with
tab_name = 'VEVO'
old_value = '23.1R1.8'
new_value = '23.2R1.15'

change_count = 0

for filename in os.listdir(directory):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(directory, filename)

        print(f"Processing file: {file_path}")

        wb = openpyxl.load_workbook(file_path)
        ws = wb[tab_name]

        version_col_index = None
        for col_num, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
            if col_cells[0].value == 'version':
                version_col_index = col_num
                break

        if version_col_index is None:
            print(f"'version' column not found in {filename}")
            continue

        for cell in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=version_col_index, max_col=version_col_index):
            for c in cell:
                print(f"Checking cell {c.coordinate}, value: {c.value}")
                if str(c.value) == old_value:
                    print(f"Changing cell {c.coordinate} in {filename} from {c.value} to {new_value}")
                    c.value = new_value
                    change_count += 1

        wb.save(file_path)

if change_count == 0:
    print("No changes made.")
else:
    print(f"Total changes made: {change_count}")
